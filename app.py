from flask import Flask, jsonify, request
from flask_sqlalchemy import SQLAlchemy
from flask_marshmallow import Marshmallow
from openpyxl import load_workbook
from flask_jwt_extended import JWTManager, jwt_required, create_access_token, get_jwt_identity

app = Flask(__name__)
app.config["SECRET_KEY"] = "your-secret-key"  # Change this to a secure secret key
app.config["JWT_SECRET_KEY"] = "your-jwt-secret-key"  # Change this to a secure JWT secret key
app.config["SQLALCHEMY_DATABASE_URI"] = 'sqlite:///db.sqllite3'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = False  # Disable token expiration for this example

db = SQLAlchemy(app)
ma = Marshmallow(app)
jwt = JWTManager(app)

class Fparser(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    age = db.Column(db.Integer, nullable=False)
    city = db.Column(db.String(200), nullable=False)

class parserschma(ma.Schema):
    class Meta:
        fields = ['id', 'name', 'age', 'city']

parser_schema = parserschma()
parser_schemas = parserschma(many=True)

# JWT Authentication
@app.route('/login', methods=['POST'])
def login():
    username = request.json.get('username', None)
    password = request.json.get('password', None)

    # Add your user authentication logic here (e.g., check username and password against the database)
    # For simplicity, let's assume you have a hardcoded user
    if username == 'user' and password == 'password':
        access_token = create_access_token(identity=username)
        return jsonify(access_token=access_token), 200
    else:
        return jsonify({"msg": "Bad username or password"}), 401

# CRUD operations with JWT authentication
@app.route('/add', methods=['POST'])
def add_data():
    if request.method == 'POST':
        excel_data = request.files['Mydata']
        Mydata = load_workbook(excel_data)
        newdata = Mydata.active
        for i in newdata.iter_rows(min_row=2, values_only=True):
            data = Fparser(name=i[0], age=i[1], city=i[2])
            db.session.add(data)
        db.session.commit()
    return "message: Data retrieved"

@app.route('/get', methods=['GET'])
@jwt_required()
def get_all_data():
    all_post = Fparser.query.all()
    result = parser_schemas.dump(all_post)
    return jsonify(result)

@app.route('/get/<int:id>', methods=['GET'])
@jwt_required()
def get_data(id):
    post = Fparser.query.filter_by(id=id).first()
    print(post)
    result = parser_schema.dump(post)
    return result

@app.route('/update/<int:id>', methods=['PUT'])
@jwt_required()
def update_data(id):
    post = Fparser.query.get(id)
    name = request.json['name']
    age = request.json['age']
    city = request.json['city']

    post.name = name
    post.age = age
    post.city = city
    db.session.commit()
    return parser_schema.jsonify(post)

@app.route('/delete/<int:id>', methods=['DELETE'])
@jwt_required()
def delete_data(id):
    post = Fparser.query.get(id)
    db.session.delete(post)
    db.session.commit()
    return parser_schema.jsonify(post)

if __name__ == '__main__':
    app.run(debug=True)
